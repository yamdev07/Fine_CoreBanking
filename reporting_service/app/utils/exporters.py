"""
Exports — Génération PDF (ReportLab) et Excel (openpyxl).
"""

import io
from decimal import Decimal

from app.core.config import settings

# ─── Helpers communs ──────────────────────────────────────────────────────────


def fmt_amount(v: Decimal | None, currency: str = "XOF") -> str:
    if v is None:
        return "-"
    formatted = f"{v:,.0f}" if currency == "XOF" else f"{v:,.2f}"
    return f"{formatted} {currency}"


def fmt_pct(v: Decimal | None) -> str:
    return f"{v:.2f}%" if v is not None else "-"


# ─── Export Excel ─────────────────────────────────────────────────────────────


def export_trial_balance_excel(report: dict) -> bytes:
    """Balance générale → fichier Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Générale"

    # En-tête
    ws.merge_cells("A1:J1")
    ws["A1"] = settings.INSTITUTION_NAME
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = report["header"]["report_title"]
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:J3")
    ws["A3"] = f"Période : {report['header']['period_start']} au {report['header']['period_end']}"
    ws["A3"].alignment = Alignment(horizontal="center")

    # En-têtes colonnes
    headers = [
        "Code",
        "Intitulé du compte",
        "Solde ouverture D",
        "Solde ouverture C",
        "Mouvement D",
        "Mouvement C",
        "Cumul D",
        "Cumul C",
        "Solde clôture D",
        "Solde clôture C",
    ]
    header_fill = PatternFill(fgColor="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Données
    row_n = 6
    current_class = None
    class_fill = PatternFill(fgColor="D6E4F0", fill_type="solid")
    alt_fill = PatternFill(fgColor="EBF5FB", fill_type="solid")

    for line in report["lines"]:
        cls = line["account_class"]
        if cls != current_class:
            current_class = cls
            cls_labels = {
                "1": "CLASSE 1 — Capitaux",
                "2": "CLASSE 2 — Actifs immobilisés",
                "3": "CLASSE 3 — Opérations interbancaires",
                "4": "CLASSE 4 — Opérations clientèle",
                "5": "CLASSE 5 — Trésorerie",
                "6": "CLASSE 6 — Charges",
                "7": "CLASSE 7 — Produits",
            }
            ws.merge_cells(f"A{row_n}:J{row_n}")
            ws[f"A{row_n}"] = cls_labels.get(cls, f"Classe {cls}")
            ws[f"A{row_n}"].font = Font(bold=True, size=10)
            ws[f"A{row_n}"].fill = class_fill
            row_n += 1

        fill = alt_fill if row_n % 2 == 0 else PatternFill()
        values = [
            line["account_code"],
            line["account_name"],
            float(line["opening_debit"]),
            float(line["opening_credit"]),
            float(line["period_debit"]),
            float(line["period_credit"]),
            float(line["cumulative_debit"]),
            float(line["cumulative_credit"]),
            float(line["closing_debit"]),
            float(line["closing_credit"]),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col > 2:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        row_n += 1

    # Totaux
    total_fill = PatternFill(fgColor="1F4E79", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    totals = [
        "TOTAUX",
        "",
        float(report["total_opening_debit"]),
        float(report["total_opening_credit"]),
        float(report["total_period_debit"]),
        float(report["total_period_credit"]),
        float(report["total_closing_debit"]),
        float(report["total_closing_credit"]),
        float(report["total_closing_debit"]),
        float(report["total_closing_credit"]),
    ]
    for col, val in enumerate(totals, start=1):
        cell = ws.cell(row=row_n, column=col, value=val)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if col > 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_dashboard_excel(report: dict) -> bytes:
    """Tableau de bord → Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau de Bord"

    ws["A1"] = settings.INSTITUTION_NAME
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Tableau de Bord au {report['as_of_date']}"
    ws["A2"].font = Font(bold=True, size=12)

    kpi_fields = [
        "kpi_encours_credits",
        "kpi_encours_epargne",
        "kpi_tresorerie",
        "kpi_produit_net_bancaire",
        "kpi_taux_impayes",
        "kpi_taux_couverture",
        "kpi_resultat_net",
        "kpi_roe",
        "kpi_roa",
        "kpi_ratio_liquidite",
        "kpi_ratio_credits_depots",
    ]

    hdr_fill = PatternFill(fgColor="1F4E79", fill_type="solid")
    for col, h in enumerate(["Indicateur", "Valeur", "Unité", "Tendance"], 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for i, field in enumerate(kpi_fields, start=5):
        kpi = report.get(field, {})
        ws.cell(row=i, column=1, value=kpi.get("label", field))
        ws.cell(row=i, column=2, value=float(kpi.get("value", 0)))
        ws.cell(row=i, column=3, value=kpi.get("unit", ""))
        ws.cell(row=i, column=4, value=kpi.get("trend", ""))
        if kpi.get("unit") == "XOF":
            ws.cell(row=i, column=2).number_format = "#,##0"
        elif kpi.get("unit") == "%":
            ws.cell(row=i, column=2).number_format = '0.00"%"'

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_journal_centralizer_excel(report: dict) -> bytes:
    """Journal centralisateur → Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Centralisateur"

    ws.merge_cells("A1:F1")
    ws["A1"] = settings.INSTITUTION_NAME
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Journal Centralisateur — "
        f"{report['header']['period_start']} au {report['header']['period_end']}"
    )
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    hdr_fill = PatternFill(fgColor="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    headers = [
        "Code journal",
        "Libellé",
        "Nb écritures",
        "Total Débit",
        "Total Crédit",
        "Équilibré",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    alt_fill = PatternFill(fgColor="EBF5FB", fill_type="solid")
    for i, line in enumerate(report["lines"], start=5):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        values = [
            line["journal_code"],
            line["journal_name"],
            line["nb_ecritures"],
            float(line["total_debit"]),
            float(line["total_credit"]),
            "OUI" if line["is_balanced"] else "NON",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = thin
            if col in (4, 5):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            if col == 6:
                cell.font = Font(
                    bold=True,
                    color="1D7A3B" if line["is_balanced"] else "C0392B",
                )

    # Ligne totaux
    row_tot = 5 + len(report["lines"])
    tot_fill = PatternFill(fgColor="1F4E79", fill_type="solid")
    tot_font = Font(bold=True, color="FFFFFF")
    totals = [
        "TOTAL",
        "",
        report["total_ecritures"],
        float(report["grand_total_debit"]),
        float(report["grand_total_credit"]),
        "OUI" if report["is_balanced"] else "NON",
    ]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row_tot, column=col, value=val)
        cell.fill = tot_fill
        cell.font = tot_font
        cell.border = thin
        if col in (4, 5):
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    for col in ("D", "E"):
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A5"

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Export PDF ───────────────────────────────────────────────────────────────


def export_bceao_pdf(report: dict) -> bytes:
    """Rapport BCEAO → PDF via ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    dark_blue = colors.HexColor("#1F4E79")
    light_blue = colors.HexColor("#D6E4F0")
    green = colors.HexColor("#1D7A3B")
    red = colors.HexColor("#C0392B")

    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        textColor=dark_blue,
        fontSize=16,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "subtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=4,
    )
    body_style = styles["Normal"]

    story = []

    # En-tête
    story.append(Paragraph(settings.INSTITUTION_NAME, title_style))
    story.append(Paragraph(report["header"]["report_title"], title_style))
    story.append(
        Paragraph(
            f"Agrément N° {report['institution_agree']} | Date d'arrêté : {report['date_arrete']}",
            subtitle_style,
        )
    )
    story.append(HRFlowable(width="100%", thickness=1, color=dark_blue))
    story.append(Spacer(1, 0.5 * cm))

    # Fonds propres nets
    story.append(
        Paragraph(
            f"<b>Fonds Propres Nets :</b> {fmt_amount(Decimal(str(report['fonds_propres_nets'])))}",
            body_style,
        )
    )
    story.append(Spacer(1, 0.4 * cm))

    # Table des ratios
    ratio_keys = [
        "ratio_solvabilite",
        "ratio_liquidite",
        "ratio_transformation",
        "ratio_division_risques",
        "ratio_couverture_risques",
    ]
    table_data = [["Code", "Libellé", "Valeur", "Norme", "Statut"]]
    for key in ratio_keys:
        r = report[key]
        conforme = r["conforme"]
        statut = "CONFORME" if conforme else "NON CONFORME"
        table_data.append(
            [
                r["code_ratio"],
                r["libelle"][:55],
                fmt_pct(Decimal(str(r["valeur"]))),
                r["norme"],
                statut,
            ]
        )

    tbl = Table(table_data, colWidths=[1.5 * cm, 9 * cm, 2.5 * cm, 2.5 * cm, 3 * cm])
    tbl_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), dark_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light_blue]),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
    )
    # Colorier la colonne Statut
    for i, key in enumerate(ratio_keys, start=1):
        r = report[key]
        color = green if r["conforme"] else red
        tbl_style.add("TEXTCOLOR", (4, i), (4, i), color)
        tbl_style.add("FONTNAME", (4, i), (4, i), "Helvetica-Bold")

    tbl.setStyle(tbl_style)
    story.append(tbl)
    story.append(Spacer(1, 0.5 * cm))

    # Synthèse
    conformes = report["ratios_conformes"]
    total = report["total_ratios"]
    color_txt = "green" if conformes == total else "red"
    story.append(
        Paragraph(
            f"<b>Synthèse :</b> <font color='{color_txt}'>"
            f"{conformes}/{total} ratios conformes</font>",
            body_style,
        )
    )

    if report.get("observations"):
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(f"<b>Observations :</b> {report['observations']}", body_style))

    story.append(Spacer(1, 0.5 * cm))
    story.append(
        Paragraph(
            f"Rapport généré le {report['header']['generated_at']} par {settings.EXPORT_AUTHOR}",
            subtitle_style,
        )
    )

    doc.build(story)
    return buf.getvalue()


def export_generic_pdf(report: dict, title: str) -> bytes:
    """PDF générique pour balance, compte de résultat, bilan."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    dark_blue = colors.HexColor("#1F4E79")

    title_s = ParagraphStyle(
        "t", parent=styles["Heading1"], textColor=dark_blue, fontSize=14, spaceAfter=4
    )
    sub_s = ParagraphStyle(
        "s", parent=styles["Normal"], fontSize=9, textColor=colors.gray, spaceAfter=4
    )
    body_s = styles["Normal"]

    story = [
        Paragraph(settings.INSTITUTION_NAME, title_s),
        Paragraph(title, title_s),
        Paragraph(
            f"Période : {report['header']['period_start']} au {report['header']['period_end']}",
            sub_s,
        ),
        HRFlowable(width="100%", thickness=1, color=dark_blue),
        Spacer(1, 0.5 * cm),
        Paragraph(
            "Ce rapport est disponible en détail via l'API (format JSON) "
            "ou en export Excel pour une analyse approfondie.",
            body_s,
        ),
        Spacer(1, 0.3 * cm),
        Paragraph(
            f"Généré le {report['header']['generated_at']}",
            sub_s,
        ),
    ]
    doc.build(story)
    return buf.getvalue()
